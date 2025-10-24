import matplotlib.pyplot as plt
import os
import streamlit as st
import japanize_matplotlib
import matplotlib.font_manager as font_manager
import pandas as pd
import numpy as np
from scipy import stats
import warnings
from typing import Optional, Dict, Any, Tuple
import requests
import json


def display_header():
    st.caption('Created by Dit-Lab.(Daiki Ito)')


def set_font():
    # font_path = '../ipaexg.ttf'  # pages フォルダから一つ上を見る
    font_path = os.path.join(os.path.dirname(__file__), '..', 'ipaexg.ttf')

    plt.rcParams['font.family'] = 'IPAexGothic'


def display_guide():
    st.markdown("""
    - [**情報探究ステップアップガイド**](https://dit-lab.notion.site/612d9665350544aa97a2a8514a03c77c?v=85ad37a3275b4717a0033516b9cfd9cc)
    - [**中の人のページ（Dit-Lab.）**](https://dit-lab.notion.site/Dit-Lab-da906d09d3cf42a19a011cf4bf25a673?pvs=4)
    """)


def display_link():
    st.header('リンク')
    st.markdown("""
    - [**中の人のページ（Dit-Lab.）**](https://dit-lab.notion.site/Dit-Lab-da906d09d3cf42a19a011cf4bf25a673?pvs=4)
    - [**進数変換学習アプリ**](https://easy-base-converter.streamlit.app)
    - [**easyRSA**](https://easy-rsa.streamlit.app/)
    - [**easyAutoML（回帰）**](https://huggingface.co/spaces/itou-daiki/pycaret_datascience_streamlit)
    - [**pkl_predict_reg**](https://huggingface.co/spaces/itou-daiki/pkl_predict_reg)
    - [**音のデータサイエンス**](https://audiovisualizationanalysis-bpeekdjwymuf6nkqcb4cqy.streamlit.app)
    - [**3D RGB Cube Visualizer**](https://boxplot-4-mysteams.streamlit.app)
    - [**上マーク角度計算補助ツール**](https://sailing-mark-angle.streamlit.app)
    - [**Factor Score Calculator**](https://factor-score-calculator.streamlit.app/)
    - [**easy Excel Merge**](https://easy-xl-merge.streamlit.app)
    - [**フィードバックはこちらまで**](https://forms.gle/G5sMYm7dNpz2FQtU9)
    - [**ソースコードはこちら（GitHub）**](https://github.com/itou-daiki/easy_stat)
    """)


def display_copyright():
    st.subheader('')
    st.write('ご意見・ご要望は→', 'https://forms.gle/G5sMYm7dNpz2FQtU9', 'まで')
    st.write('')
    st.subheader('© 2022-2025 Dit-Lab.(Daiki Ito). All Rights Reserved.')
    st.write('easyStat: Open Source for Ubiquitous Statistics')
    st.write('Democratizing data, everywhere.')
    st.write('')


def display_special_thanks():
    st.subheader('In collaboration with our esteemed contributors:')
    st.write('・Toshiyuki')
    st.write('With heartfelt appreciation for their dedication and support.')


# ==========================================
# 統計初学者向け学習支援機能
# ==========================================

class StatisticalLearningAssistant:
    """統計初学者向けの学習支援クラス"""
    
    def __init__(self):
        self.learning_levels = {
            'beginner': '初級者',
            'intermediate': '中級者', 
            'advanced': '上級者'
        }
    
    def show_concept_explanation(self, concept_key: str, level: str = 'beginner'):
        """統計概念の説明を表示"""
        explanations = {
            'correlation': {
                'beginner': """
                📊 **相関分析とは？**
                
                相関分析は、2つの変数がどれくらい関係しているかを調べる分析方法です。
                
                **相関係数の読み方：**
                - 1に近い：強い正の相関（一方が増えると他方も増える）
                - 0に近い：相関なし（関係が薄い）
                - -1に近い：強い負の相関（一方が増えると他方は減る）
                
                **例：** 勉強時間と成績、身長と体重など
                """,
                'intermediate': """
                📊 **相関分析の詳細**
                
                ピアソンの相関係数（r）は線形関係の強さを測定します。
                
                **解釈の目安：**
                - |r| ≥ 0.7：強い相関
                - 0.3 ≤ |r| < 0.7：中程度の相関
                - |r| < 0.3：弱い相関
                
                **注意点：** 相関関係≠因果関係
                """,
                'advanced': """
                📊 **相関分析の統計的詳細**
                
                - ピアソンの積率相関係数：r = Σ[(xi-x̄)(yi-ȳ)] / √[Σ(xi-x̄)²Σ(yi-ȳ)²]
                - 前提条件：正規分布、線形関係、等分散性
                - 有意性検定：t = r√(n-2)/√(1-r²)
                """
            },
            'ttest': {
                'beginner': """
                📊 **t検定とは？**
                
                2つのグループの平均値に違いがあるかを調べる検定です。
                
                **種類：**
                - 対応なし：異なる人たちを比較（例：男性vs女性の身長）
                - 対応あり：同じ人の前後を比較（例：薬の服用前後）
                
                **p値が0.05未満なら「有意な差がある」と判断します**
                """,
                'intermediate': """
                📊 **t検定の詳細**
                
                **前提条件：**
                - データが正規分布に従う
                - 対応なしの場合：等分散性
                - 独立性
                
                **効果量（Cohen's d）：**
                - 0.2：小さい効果
                - 0.5：中程度の効果  
                - 0.8：大きい効果
                """,
                'advanced': """
                📊 **t検定の統計的詳細**
                
                - 対応なし：t = (x̄₁-x̄₂) / SE_diff
                - 対応あり：t = d̄ / (sd/√n)
                - Welchのt検定：等分散性を仮定しない
                - 自由度の調整が重要
                """
            }
        }
        
        if concept_key in explanations and level in explanations[concept_key]:
            st.info(explanations[concept_key][level])
    
    def check_learning_progress(self, analysis_type: str):
        """学習進捗をチェック"""
        if 'learning_progress' not in st.session_state:
            st.session_state.learning_progress = set()
        
        st.session_state.learning_progress.add(analysis_type)
        
        progress_count = len(st.session_state.learning_progress)
        total_analyses = 14  # 総分析数
        
        st.sidebar.success(f"📚 学習進捗: {progress_count}/{total_analyses} 完了")
        
        if progress_count >= total_analyses:
            st.balloons()
            st.success("🎉 すべての分析を体験しました！統計マスターですね！")


# ==========================================
# エラーハンドリング・データ検証フレームワーク
# ==========================================

class StatisticalValidator:
    """統計分析のデータ検証クラス"""
    
    @staticmethod
    def safe_file_load(uploaded_file) -> Optional[pd.DataFrame]:
        """安全なファイル読み込み"""
        try:
            if uploaded_file is None:
                return None
                
            file_extension = uploaded_file.name.split('.')[-1].lower()
            
            if file_extension == 'csv':
                # CSVの場合、エンコーディングを自動判定
                try:
                    df = pd.read_csv(uploaded_file, encoding='utf-8')
                except UnicodeDecodeError:
                    uploaded_file.seek(0)  # ファイルポインタをリセット
                    df = pd.read_csv(uploaded_file, encoding='shift_jis')
            elif file_extension in ['xlsx', 'xls']:
                df = pd.read_excel(uploaded_file)
            else:
                st.error("⚠️ 対応していないファイル形式です。CSV、Excel(.xlsx/.xls)ファイルをアップロードしてください。")
                return None
            
            # 基本的なデータ検証
            if df.empty:
                st.error("⚠️ アップロードされたファイルにデータが含まれていません。")
                return None
            
            # 列名の日本語対応確認
            if any(col.startswith('Unnamed:') for col in df.columns):
                st.warning("⚠️ 列名が正しく読み込まれていない可能性があります。ファイルの1行目に列名が含まれているか確認してください。")
            
            return df
            
        except Exception as e:
            st.error(f"⚠️ ファイルの読み込み中にエラーが発生しました: {str(e)}")
            st.info("💡 **解決のヒント:**\n- ファイルが破損していないか確認\n- Excel形式の場合は.xlsxで保存\n- CSVの場合は文字コード(UTF-8)を確認")
            return None
    
    @staticmethod
    def validate_sample_size(data: pd.DataFrame, min_size: int = 3, analysis_type: str = "分析") -> bool:
        """サンプルサイズの検証"""
        if len(data) < min_size:
            st.error(f"⚠️ {analysis_type}には最低{min_size}件のデータが必要です。現在のデータ件数: {len(data)}件")
            st.info("💡 **解決方法:** より多くのデータを収集するか、他の分析手法を検討してください。")
            return False
        return True
    
    @staticmethod
    def check_missing_values(data: pd.DataFrame, columns: list) -> Dict[str, Any]:
        """欠損値のチェック"""
        missing_info = {}
        for col in columns:
            if col in data.columns:
                missing_count = data[col].isnull().sum()
                missing_info[col] = missing_count
                
                if missing_count > 0:
                    st.warning(f"⚠️ 変数「{col}」に{missing_count}件の欠損値があります。")
        
        total_missing = sum(missing_info.values())
        if total_missing > 0:
            st.info("💡 **欠損値の対処法:**\n- データクレンジングページで欠損値を処理\n- 欠損値のある行を除外\n- 平均値や中央値で補完")
        
        return missing_info
    
    @staticmethod
    def validate_data_types(data: pd.DataFrame, columns: list, expected_type: str = 'numeric') -> bool:
        """データ型の検証"""
        valid = True
        for col in columns:
            if col in data.columns:
                if expected_type == 'numeric':
                    if not pd.api.types.is_numeric_dtype(data[col]):
                        st.error(f"⚠️ 変数「{col}」は数値型である必要があります。現在の型: {data[col].dtype}")
                        st.info("💡 **解決方法:** データクレンジングページで数値型に変換してください。")
                        valid = False
        return valid
    
    @staticmethod
    def check_normality(data: pd.Series, alpha: float = 0.05) -> Dict[str, Any]:
        """正規性の検定"""
        if len(data.dropna()) < 3:
            return {'test': 'insufficient_data', 'p_value': None, 'is_normal': False}
        
        # Shapiro-Wilk検定（サンプルサイズが小さい場合）
        if len(data.dropna()) <= 50:
            stat, p_value = stats.shapiro(data.dropna())
            test_name = 'Shapiro-Wilk'
        else:
            # Kolmogorov-Smirnov検定（サンプルサイズが大きい場合）
            stat, p_value = stats.kstest(data.dropna(), 'norm')
            test_name = 'Kolmogorov-Smirnov'
        
        is_normal = p_value > alpha
        return {
            'test': test_name,
            'statistic': stat,
            'p_value': p_value,
            'is_normal': is_normal
        }
    
    @staticmethod
    def check_equal_variances(group1: pd.Series, group2: pd.Series, alpha: float = 0.05) -> Dict[str, Any]:
        """等分散性の検定（Levene検定）"""
        try:
            stat, p_value = stats.levene(group1.dropna(), group2.dropna())
            equal_variances = p_value > alpha
            return {
                'statistic': stat,
                'p_value': p_value,
                'equal_variances': equal_variances
            }
        except Exception as e:
            return {'error': str(e), 'equal_variances': False}


# ==========================================
# 結果解釈支援機能
# ==========================================

class ResultInterpreter:
    """統計結果の解釈支援クラス"""
    
    @staticmethod
    def interpret_correlation(r: float, p_value: float, alpha: float = 0.05) -> str:
        """相関係数の解釈"""
        # 相関の強さ
        abs_r = abs(r)
        if abs_r >= 0.7:
            strength = "強い"
        elif abs_r >= 0.3:
            strength = "中程度の"
        else:
            strength = "弱い"
        
        # 相関の方向
        direction = "正の" if r > 0 else "負の"
        
        # 有意性
        significance = "統計的に有意" if p_value < alpha else "統計的に有意ではない"
        
        interpretation = f"""
        📊 **結果の解釈**
        
        - **相関の強さ**: {strength}相関 (r = {r:.3f})
        - **相関の方向**: {direction}相関
        - **統計的有意性**: {significance} (p = {p_value:.3f})
        
        **実際の意味**:
        """
        
        if abs_r >= 0.7 and p_value < alpha:
            interpretation += "2つの変数には強い関係があります。一方が変化すると他方も予測可能な形で変化する傾向があります。"
        elif abs_r >= 0.3 and p_value < alpha:
            interpretation += "2つの変数にはある程度の関係があります。完全ではありませんが、関連性が認められます。"
        elif p_value >= alpha:
            interpretation += "統計的に有意な関係は認められませんでした。偶然による結果の可能性があります。"
        else:
            interpretation += "関係はあるものの弱く、実際の予測や判断には注意が必要です。"
        
        return interpretation
    
    @staticmethod
    def interpret_ttest(t_stat: float, p_value: float, effect_size: float, alpha: float = 0.05) -> str:
        """t検定結果の解釈"""
        significance = "統計的に有意" if p_value < alpha else "統計的に有意ではない"
        
        # 効果量の解釈
        if abs(effect_size) >= 0.8:
            effect_interpretation = "大きい効果"
        elif abs(effect_size) >= 0.5:
            effect_interpretation = "中程度の効果"
        elif abs(effect_size) >= 0.2:
            effect_interpretation = "小さい効果"
        else:
            effect_interpretation = "効果はほとんどない"
        
        interpretation = f"""
        📊 **結果の解釈**
        
        - **統計的有意性**: {significance} (p = {p_value:.3f})
        - **効果量**: {effect_interpretation} (Cohen's d = {effect_size:.3f})
        - **t統計量**: {t_stat:.3f}
        
        **結論**:
        """
        
        if p_value < alpha:
            interpretation += f"2つのグループ間に{effect_interpretation}な差があることが統計的に確認されました。"
        else:
            interpretation += "2つのグループ間に統計的に有意な差は認められませんでした。"
        
        return interpretation


# ==========================================
# インタラクティブ学習ガイド
# ==========================================

def show_interactive_guide(analysis_type: str):
    """インタラクティブな学習ガイドを表示"""
    guides = {
        'correlation': {
            'title': '🔍 相関分析学習ガイド',
            'steps': [
                "1️⃣ データをアップロードしましょう",
                "2️⃣ 分析したい2つの変数を選択します",
                "3️⃣ 散布図で視覚的に関係を確認",
                "4️⃣ 相関係数を計算・解釈",
                "5️⃣ 結果をレポートにまとめる"
            ],
            'tips': [
                "💡 まずは散布図で直線的な関係があるか確認",
                "💡 外れ値がないかチェック",
                "💡 相関≠因果関係を忘れずに"
            ]
        },
        'ttest': {
            'title': '📊 t検定学習ガイド',
            'steps': [
                "1️⃣ 比較したいグループを明確にする",
                "2️⃣ データの正規性を確認",
                "3️⃣ 等分散性をチェック（対応なしの場合）",
                "4️⃣ 適切なt検定を実行",
                "5️⃣ 効果量も含めて解釈"
            ],
            'tips': [
                "💡 対応あり・なしの選択が重要",
                "💡 p値だけでなく効果量も確認",
                "💡 実際的な意味を考える"
            ]
        }
    }
    
    if analysis_type in guides:
        guide = guides[analysis_type]
        
        with st.expander(f"{guide['title']} - クリックして展開"):
            st.markdown("### 📋 学習ステップ")
            for step in guide['steps']:
                st.markdown(step)
            
            st.markdown("### 💡 重要なポイント")
            for tip in guide['tips']:
                st.markdown(tip)
            
            # 学習チェックリスト
            st.markdown("### ✅ 学習チェックリスト")
            for i, step in enumerate(guide['steps']):
                completed = st.checkbox(f"完了: {step}", key=f"checklist_{analysis_type}_{i}")


def show_beginner_tips():
    """初学者向けのヒントを表示"""
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🎓 統計学習のコツ")
    
    tips = [
        "📊 まずはグラフで視覚化",
        "🔢 数値だけでなく意味も考える", 
        "❓ 「なぜ？」を常に問いかける",
        "📈 複数の分析を組み合わせる",
        "📝 結果を文章で説明してみる"
    ]
    
    for tip in tips:
        st.sidebar.markdown(f"- {tip}")


def create_learning_dashboard():
    """学習ダッシュボードを作成"""
    if 'learning_progress' not in st.session_state:
        st.session_state.learning_progress = set()
    
    progress = st.session_state.learning_progress
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📚 学習進捗ダッシュボード")
    
    analyses = [
        "データクレンジング", "EDA", "相関分析", "カイ二乗検定",
        "t検定（対応なし）", "t検定（対応あり）", "一要因分散分析（対応なし）",
        "一要因分散分析（対応あり）", "二要因分散分析", "二要因混合分散分析",
        "単回帰分析", "重回帰分析", "因子分析", "テキストマイニング"
    ]
    
    completed_count = len(progress)
    total_count = len(analyses)
    progress_percentage = (completed_count / total_count) * 100
    
    st.sidebar.progress(progress_percentage / 100)
    st.sidebar.markdown(f"**進捗: {completed_count}/{total_count} ({progress_percentage:.1f}%)**")
    
    if completed_count == total_count:
        st.sidebar.success("🏆 全分析制覇！")
    elif completed_count >= total_count // 2:
        st.sidebar.info("📈 中級者レベル達成！")
    elif completed_count >= 3:
        st.sidebar.info("🌱 順調に学習中！")


# ==========================================
# 生成AI統計解釈支援機能
# ==========================================

class AIStatisticalInterpreter:
    """生成AIによる統計解釈支援クラス"""

    @staticmethod
    def call_gemini_api(api_key: str, prompt: str) -> str:
        """Gemini 2.0 Flash APIを呼び出す関数"""
        if not api_key:
            return "APIキーが設定されていません。"

        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent"
        headers = {
            "Content-Type": "application/json",
        }

        data = {
            "contents": [{
                "parts": [{
                    "text": prompt
                }]
            }],
            "generationConfig": {
                "temperature": 0.3,
                "maxOutputTokens": 2048,
            }
        }

        try:
            response = requests.post(f"{url}?key={api_key}", headers=headers, json=data)
            if response.status_code == 200:
                result = response.json()
                if 'candidates' in result and len(result['candidates']) > 0:
                    return result['candidates'][0]['content']['parts'][0]['text']
                else:
                    return "APIからの応答が予期しない形式です。"
            else:
                return f"APIエラー: {response.status_code} - {response.text}"
        except Exception as e:
            return f"エラーが発生しました: {str(e)}"

    @staticmethod
    def setup_ai_sidebar() -> Tuple[str, bool]:
        """AI機能のサイドバー設定"""
        st.sidebar.subheader("🤖 AI統計解釈機能")
        st.sidebar.write("Gemini 2.0 Flash APIを使用して統計結果を自動解釈します")
        gemini_api_key = st.sidebar.text_input(
            "Gemini APIキーを入力してください",
            type="password",
            help="Google AI Studio (https://aistudio.google.com/) でAPIキーを取得できます"
        )
        enable_ai_interpretation = st.sidebar.checkbox(
            "AI解釈機能を有効にする",
            disabled=not gemini_api_key
        )

        if gemini_api_key and enable_ai_interpretation:
            st.sidebar.success("✅ AI解釈機能が有効になりました")
        elif enable_ai_interpretation and not gemini_api_key:
            st.sidebar.error("❌ APIキーを入力してください")

        return gemini_api_key, enable_ai_interpretation

    @staticmethod
    def create_correlation_interpretation_prompt(correlation_results: Dict[str, Any]) -> str:
        """相関分析の解釈プロンプト"""
        r = correlation_results.get('correlation', 0)
        p_value = correlation_results.get('p_value', 1)
        var1 = correlation_results.get('var1', '変数1')
        var2 = correlation_results.get('var2', '変数2')
        n = correlation_results.get('sample_size', 0)

        prompt = f"""
あなたは統計分析の専門家です。以下の相関分析の結果を詳細に解釈・考察してください。

【分析結果】
- 変数1: {var1}
- 変数2: {var2}
- サンプルサイズ: {n}
- 相関係数 (r): {r:.4f}
- p値: {p_value:.4f}

【解釈・考察してほしい内容】
1. 相関係数の値から見た関係の強さと方向性
   - 数値の具体的な意味
   - 相関の強度の評価（弱い/中程度/強い）

2. 統計的有意性の判断
   - p値の解釈
   - 帰無仮説と対立仮説の判断

3. 実用的な解釈
   - ビジネスや研究文脈での意味
   - 実際的な影響度の評価

4. 注意点と限界
   - 相関と因果の区別
   - 第三の変数の可能性
   - サンプルサイズの妥当性

5. 今後の分析提案
   - 追加で実施すべき分析
   - 深掘りすべきポイント

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def create_chi_square_interpretation_prompt(chi_square_results: Dict[str, Any]) -> str:
        """カイ二乗検定の解釈プロンプト"""
        chi2 = chi_square_results.get('chi2', 0)
        p_value = chi_square_results.get('p_value', 1)
        dof = chi_square_results.get('dof', 0)
        var1 = chi_square_results.get('var1', '変数1')
        var2 = chi_square_results.get('var2', '変数2')
        crosstab = chi_square_results.get('crosstab', None)
        expected = chi_square_results.get('expected', None)

        crosstab_str = crosstab.to_string() if crosstab is not None else "データなし"
        expected_str = expected.to_string() if expected is not None else "データなし"

        prompt = f"""
あなたは統計分析の専門家です。以下のカイ二乗検定の結果を詳細に解釈・考察してください。

【分析結果】
- 変数1: {var1}
- 変数2: {var2}
- カイ二乗統計量: {chi2:.4f}
- 自由度: {dof}
- p値: {p_value:.4f}

【観測度数（クロス表）】
{crosstab_str}

【期待度数】
{expected_str}

【解釈・考察してほしい内容】
1. クロス表の読み取り
   - 観測度数のパターン
   - 顕著な偏りの特定

2. カイ二乗統計量とp値の解釈
   - 統計的有意性の判断
   - 実際的な関連の強さ

3. 期待度数との差異分析
   - どのセルで大きな差があるか
   - その実際的な意味

4. 変数間の関連性の解釈
   - 独立性の検定結果
   - 関連のパターン

5. 実用的な示唆
   - ビジネス・研究での活用
   - 意思決定への応用

6. 注意点と限界
   - 期待度数の妥当性
   - 因果関係の解釈制限

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def create_ttest_interpretation_prompt(ttest_results: Dict[str, Any]) -> str:
        """t検定の解釈プロンプト"""
        t_stat = ttest_results.get('t_statistic', 0)
        p_value = ttest_results.get('p_value', 1)
        dof = ttest_results.get('dof', 0)
        mean1 = ttest_results.get('mean1', 0)
        mean2 = ttest_results.get('mean2', 0)
        std1 = ttest_results.get('std1', 0)
        std2 = ttest_results.get('std2', 0)
        n1 = ttest_results.get('n1', 0)
        n2 = ttest_results.get('n2', 0)
        effect_size = ttest_results.get('effect_size', 0)
        test_type = ttest_results.get('test_type', 't検定')
        group1_name = ttest_results.get('group1_name', 'グループ1')
        group2_name = ttest_results.get('group2_name', 'グループ2')

        prompt = f"""
あなたは統計分析の専門家です。以下の{test_type}の結果を詳細に解釈・考察してください。

【分析結果】
- 検定タイプ: {test_type}
- グループ1 ({group1_name}): 平均={mean1:.4f}, 標準偏差={std1:.4f}, n={n1}
- グループ2 ({group2_name}): 平均={mean2:.4f}, 標準偏差={std2:.4f}, n={n2}
- t統計量: {t_stat:.4f}
- 自由度: {dof}
- p値: {p_value:.4f}
- 効果量 (Cohen's d): {effect_size:.4f}

【解釈・考察してほしい内容】
1. 記述統計の比較
   - 平均値の差の実際的な大きさ
   - 標準偏差から見るばらつき

2. t統計量とp値の解釈
   - 統計的有意性の判断
   - 帰無仮説の採択/棄却

3. 効果量の評価
   - Cohen's dの解釈（小/中/大）
   - 実際的な意味での差の大きさ

4. グループ間の差の実用的解釈
   - ビジネス・研究での意味
   - 意思決定への示唆

5. サンプルサイズの妥当性
   - 統計的検出力の評価
   - 結果の信頼性

6. 注意点と限界
   - 前提条件の確認
   - 因果推論の制限
   - 追加分析の提案

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def create_anova_interpretation_prompt(anova_results: Dict[str, Any]) -> str:
        """分散分析の解釈プロンプト"""
        f_stat = anova_results.get('f_statistic', 0)
        p_value = anova_results.get('p_value', 1)
        df_between = anova_results.get('df_between', 0)
        df_within = anova_results.get('df_within', 0)
        group_means = anova_results.get('group_means', {})
        eta_squared = anova_results.get('eta_squared', 0)
        analysis_type = anova_results.get('analysis_type', '分散分析')

        means_str = "\n".join([f"- {group}: 平均={mean:.4f}" for group, mean in group_means.items()])

        prompt = f"""
あなたは統計分析の専門家です。以下の{analysis_type}の結果を詳細に解釈・考察してください。

【分析結果】
- 分析タイプ: {analysis_type}
- F統計量: {f_stat:.4f}
- 自由度: 群間={df_between}, 群内={df_within}
- p値: {p_value:.4f}
- 効果量 (η²): {eta_squared:.4f}

【各グループの平均値】
{means_str}

【解釈・考察してほしい内容】
1. 記述統計の比較
   - 各グループの平均値の傾向
   - 最大・最小の差

2. F統計量とp値の解釈
   - 統計的有意性の判断
   - グループ間の差の存在

3. 効果量の評価
   - η²の解釈
   - 実際的な差の大きさ

4. 多重比較の必要性
   - どのグループ間で差があるか
   - 事後検定の推奨

5. 実用的な示唆
   - ビジネス・研究での意味
   - 最適なグループの特定

6. 注意点と限界
   - 前提条件の確認
   - Type Iエラーの制御
   - 追加分析の提案

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def create_regression_interpretation_prompt(regression_results: Dict[str, Any]) -> str:
        """回帰分析の解釈プロンプト"""
        r_squared = regression_results.get('r_squared', 0)
        adj_r_squared = regression_results.get('adj_r_squared', 0)
        f_stat = regression_results.get('f_statistic', 0)
        f_pvalue = regression_results.get('f_pvalue', 1)
        coefficients = regression_results.get('coefficients', {})
        
        coef_str = "\n".join([f"- {var}: 係数={coef['coef']:.4f}, p値={coef['pvalue']:.4f}" 
                             for var, coef in coefficients.items()])
        
        prompt = f"""
あなたは統計分析の専門家です。以下の回帰分析の結果を詳細に解釈・考察してください。

【分析結果】
- 決定係数 (R²): {r_squared:.4f}
- 調整済みR²: {adj_r_squared:.4f}
- F統計量: {f_stat:.4f}
- F検定p値: {f_pvalue:.4f}

【回帰係数】
{coef_str}

【解釈・考察してほしい内容】
1. モデル全体の適合度
   - R²と調整済みR²の解釈
   - モデルの説明力の評価

2. モデルの統計的有意性
   - F検定の結果
   - モデル全体の有効性

3. 各説明変数の影響
   - 有意な変数の特定
   - 係数の大きさと方向性の解釈
   - 各変数の実際的な影響度

4. 実用的な示唆
   - ビジネス・研究での活用
   - 予測や意思決定への応用

5. 注意点と限界
   - 多重共線性の可能性
   - モデルの前提条件
   - 改善の余地

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def create_factor_analysis_interpretation_prompt(factor_results: Dict[str, Any]) -> str:
        """因子分析の解釈プロンプト"""
        n_factors = factor_results.get('n_factors', 0)
        variance_explained = factor_results.get('variance_explained', [])
        cumulative_variance = factor_results.get('cumulative_variance', [])
        loadings = factor_results.get('loadings', {})
        
        variance_str = "\n".join([f"- 因子{i+1}: {var:.2f}%" 
                                 for i, var in enumerate(variance_explained)])
        
        prompt = f"""
あなたは統計分析の専門家です。以下の因子分析の結果を詳細に解釈・考察してください。

【分析結果】
- 抽出因子数: {n_factors}
- 累積寄与率: {cumulative_variance[-1] if cumulative_variance else 0:.2f}%

【各因子の寄与率】
{variance_str}

【解釈・考察してほしい内容】
1. 因子数の妥当性
   - 抽出された因子数の適切性
   - 累積寄与率の評価

2. 各因子の特徴
   - 因子負荷量のパターン
   - 各因子が表す潜在的な概念

3. 因子の解釈可能性
   - 因子の命名提案
   - ビジネス・研究での意味

4. データの構造理解
   - 変数間の潜在的関係
   - データの次元削減効果

5. 実用的な活用方法
   - 因子得点の利用
   - 後続分析への応用

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def create_pca_interpretation_prompt(pca_results: Dict[str, Any]) -> str:
        """主成分分析の解釈プロンプト"""
        n_components = pca_results.get('n_components', 0)
        variance_explained = pca_results.get('variance_explained', [])
        cumulative_variance = pca_results.get('cumulative_variance', [])
        
        variance_str = "\n".join([f"- 第{i+1}主成分: {var:.2f}%" 
                                 for i, var in enumerate(variance_explained)])
        
        prompt = f"""
あなたは統計分析の専門家です。以下の主成分分析の結果を詳細に解釈・考察してください。

【分析結果】
- 主成分数: {n_components}
- 累積寄与率: {cumulative_variance[-1] if cumulative_variance else 0:.2f}%

【各主成分の寄与率】
{variance_str}

【解釈・考察してほしい内容】
1. 主成分数の妥当性
   - 選択された主成分数の適切性
   - 累積寄与率の評価
   - 情報損失の程度

2. 各主成分の特徴
   - 寄与率の分布パターン
   - 主要な主成分の意味

3. 次元削減の効果
   - データの圧縮度合い
   - 元の情報の保持状況

4. データの構造理解
   - 変数間の関係性
   - データの特徴抽出

5. 実用的な活用方法
   - 可視化への応用
   - 機械学習の前処理
   - ノイズ除去効果

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def create_eda_interpretation_prompt(eda_results: Dict[str, Any]) -> str:
        """探索的データ分析の解釈プロンプト"""
        variable_name = eda_results.get('variable_name', '変数')
        mean = eda_results.get('mean', 0)
        median = eda_results.get('median', 0)
        std = eda_results.get('std', 0)
        min_val = eda_results.get('min', 0)
        max_val = eda_results.get('max', 0)
        q1 = eda_results.get('q1', 0)
        q3 = eda_results.get('q3', 0)
        skewness = eda_results.get('skewness', 0)
        kurtosis = eda_results.get('kurtosis', 0)
        
        prompt = f"""
あなたは統計分析の専門家です。以下の探索的データ分析（EDA）の結果を詳細に解釈・考察してください。

【分析対象】
- 変数名: {variable_name}

【記述統計量】
- 平均値: {mean:.4f}
- 中央値: {median:.4f}
- 標準偏差: {std:.4f}
- 最小値: {min_val:.4f}
- 最大値: {max_val:.4f}
- 第1四分位数: {q1:.4f}
- 第3四分位数: {q3:.4f}
- 歪度: {skewness:.4f}
- 尖度: {kurtosis:.4f}

【解釈・考察してほしい内容】
1. データの中心傾向
   - 平均値と中央値の関係
   - 代表値として適切な指標

2. データのばらつき
   - 標準偏差の解釈
   - データの散らばり具合

3. データの分布形状
   - 歪度からわかる偏り
   - 尖度からわかる裾の重さ
   - 正規性の評価

4. 外れ値の可能性
   - 最小値・最大値の妥当性
   - 異常値の存在可能性

5. データの特徴まとめ
   - 全体的な傾向
   - 注意すべきポイント
   - 後続分析への示唆

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def create_text_mining_interpretation_prompt(text_results: Dict[str, Any]) -> str:
        """テキストマイニングの解釈プロンプト"""
        top_words = text_results.get('top_words', [])
        n_documents = text_results.get('n_documents', 0)
        n_unique_words = text_results.get('n_unique_words', 0)
        
        words_str = "\n".join([f"- {word}: {count}回" 
                              for word, count in top_words[:20]])
        
        prompt = f"""
あなたはテキスト分析の専門家です。以下のテキストマイニングの結果を詳細に解釈・考察してください。

【分析結果】
- 文書数: {n_documents}
- ユニーク単語数: {n_unique_words}

【頻出単語トップ20】
{words_str}

【解釈・考察してほしい内容】
1. 全体的なテーマ
   - 頻出単語から読み取れる主要テーマ
   - 文書全体のトピック

2. 単語の出現パターン
   - 特に目立つキーワード
   - 単語間の関連性の推測

3. 内容の特徴
   - ポジティブ/ネガティブな傾向
   - 専門性や抽象度

4. ビジネス・研究への示唆
   - テキストから得られる洞察
   - 実務的な活用方法

5. 追加分析の提案
   - 深掘りすべきポイント
   - 有効な分析手法

統計の専門知識がない人にも分かりやすく、実践的な解釈を日本語で提供してください。
"""
        return prompt

    @staticmethod
    def display_ai_interpretation(
        api_key: str,
        enabled: bool,
        results: Dict[str, Any],
        analysis_type: str,
        key_prefix: str = "ai_interp"
    ):
        """AI解釈を表示する共通関数"""
        if not enabled or not api_key:
            return

        st.subheader(f"🤖 AI統計解釈")

        interpretation_key = f"{key_prefix}_interpretation"

        # 解釈ボタン
        if st.button(f"統計結果を解釈する", key=f"{key_prefix}_button"):
            with st.spinner("AIが統計結果を分析中..."):
                # 分析タイプに応じたプロンプトを作成
                if analysis_type == 'correlation':
                    prompt = AIStatisticalInterpreter.create_correlation_interpretation_prompt(results)
                elif analysis_type == 'chi_square':
                    prompt = AIStatisticalInterpreter.create_chi_square_interpretation_prompt(results)
                elif analysis_type == 'ttest':
                    prompt = AIStatisticalInterpreter.create_ttest_interpretation_prompt(results)
                elif analysis_type == 'anova':
                    prompt = AIStatisticalInterpreter.create_anova_interpretation_prompt(results)
                elif analysis_type == 'regression':
                    prompt = AIStatisticalInterpreter.create_regression_interpretation_prompt(results)
                elif analysis_type == 'factor_analysis':
                    prompt = AIStatisticalInterpreter.create_factor_analysis_interpretation_prompt(results)
                elif analysis_type == 'pca':
                    prompt = AIStatisticalInterpreter.create_pca_interpretation_prompt(results)
                elif analysis_type == 'eda':
                    prompt = AIStatisticalInterpreter.create_eda_interpretation_prompt(results)
                elif analysis_type == 'text_mining':
                    prompt = AIStatisticalInterpreter.create_text_mining_interpretation_prompt(results)
                else:
                    st.error("未対応の分析タイプです。")
                    return

                # API呼び出し
                interpretation = AIStatisticalInterpreter.call_gemini_api(api_key, prompt)

                # 結果をセッション状態に保存
                st.session_state[interpretation_key] = interpretation

        # 解釈結果がある場合は常に表示
        if interpretation_key in st.session_state:
            st.markdown("### 📊 統計解釈結果")
            st.write(st.session_state[interpretation_key])

            # 解釈をクリアするボタン
            col1, col2 = st.columns([1, 1])
            with col2:
                if st.button(f"解釈をクリア", key=f"{key_prefix}_clear"):
                    del st.session_state[interpretation_key]
                    st.rerun()

# ==========================================
# グラフExport機能
# ==========================================

def export_plotly_to_excel(fig, filename="graph.xlsx", sheet_name="グラフ"):
    """
    PlotlyグラフをExcelネイティブグラフとしてExcelファイルに変換する
    
    Parameters:
    -----------
    fig : plotly.graph_objects.Figure
        変換するPlotlyグラフ
    filename : str
        保存するファイル名
    sheet_name : str
        Excelシート名
        
    Returns:
    --------
    bytes
        Excelファイルのバイナリデータ
    """
    import io
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, ScatterChart, LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    
    # Excelワークブックを作成
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # グラフタイトルを取得
    graph_title = ""
    if fig.layout.title and fig.layout.title.text:
        graph_title = fig.layout.title.text
    
    # Plotlyグラフからデータを抽出
    data_dict = {}
    categories = []
    series_names = []
    
    # グラフの種類を判定
    graph_type = None
    if fig.data:
        first_trace = fig.data[0]
        if hasattr(first_trace, 'type'):
            graph_type = first_trace.type
    
    # データを抽出してワークシートに書き込む
    if graph_type == 'bar':
        # 棒グラフの場合
        # X軸のカテゴリを取得（ticktextを優先的に使用）
        categories = []

        # まずticktextを確認
        try:
            if hasattr(fig.layout, 'xaxis') and fig.layout.xaxis is not None:
                if hasattr(fig.layout.xaxis, 'ticktext') and fig.layout.xaxis.ticktext is not None:
                    ticktext = fig.layout.xaxis.ticktext
                    # タプルまたはリストをリストに変換
                    if isinstance(ticktext, (list, tuple)) and len(ticktext) > 0:
                        categories = [str(t) for t in ticktext]
        except Exception as e:
            pass

        # ticktextが取得できなかった場合のみ、xの値を使用
        if not categories:
            try:
                if hasattr(fig.data[0], 'x') and fig.data[0].x is not None:
                    x_data = fig.data[0].x
                    if isinstance(x_data, (list, tuple)) and len(x_data) > 0:
                        categories = [str(x) for x in x_data]
            except Exception as e:
                categories = ["カテゴリ1", "カテゴリ2"]  # フォールバック
        
        # データを収集
        for trace in fig.data:
            if hasattr(trace, 'name') and trace.name:
                series_name = trace.name
            else:
                series_name = "系列1"
            
            series_names.append(series_name)
            
            if hasattr(trace, 'y') and trace.y is not None:
                data_dict[series_name] = list(trace.y)
        
        # データをワークシートに書き込む（グラフタイトルを1行目の3列目に配置）
        ws.cell(row=1, column=1, value="カテゴリ")
        col_idx = 2
        for series_name in series_names:
            ws.cell(row=1, column=col_idx, value=series_name)
            col_idx += 1
        # グラフタイトルを1行目の最後の列の次に配置
        ws.cell(row=1, column=col_idx, value=graph_title)
        
        for row_idx, category in enumerate(categories, start=2):
            ws.cell(row=row_idx, column=1, value=category)
            col_idx = 2
            for series_name in series_names:
                if row_idx - 2 < len(data_dict[series_name]):
                    ws.cell(row=row_idx, column=col_idx, value=data_dict[series_name][row_idx - 2])
                col_idx += 1
        
        # 棒グラフを作成
        chart = BarChart()
        chart.type = "col"  # 縦棒グラフ
        chart.style = None  # プレーンな設定
        if graph_title:
            chart.title = graph_title
        chart.y_axis.title = fig.layout.yaxis.title.text if fig.layout.yaxis and fig.layout.yaxis.title else ""
        chart.x_axis.title = fig.layout.xaxis.title.text if fig.layout.xaxis and fig.layout.xaxis.title else ""
        
        # データ範囲を設定
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(categories) + 1, max_col=len(series_names) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(categories) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # グラフのサイズを設定
        chart.width = 15  # 幅（cm単位）
        chart.height = 10  # 高さ（cm単位）
        
        # 凡例の設定
        if len(series_names) > 1:
            chart.legend.position = 'r'  # 凡例を右側に配置
        else:
            chart.legend = None  # 凡例を非表示（系列が1つの場合は不要）
        
        # グラフの配置（データの下に配置）
        ws.add_chart(chart, "A" + str(len(categories) + 3))
        
    elif graph_type == 'scatter':
        # 散布図の場合
        # グラフタイトルを1行目の最初に配置
        ws.cell(row=1, column=1, value=graph_title)
        
        col_offset = 0
        for trace_idx, trace in enumerate(fig.data):
            if hasattr(trace, 'name') and trace.name:
                series_name = trace.name
            else:
                series_name = f"系列{trace_idx + 1}"
            
            series_names.append(series_name)
            
            if hasattr(trace, 'x') and trace.x is not None:
                x_data = list(trace.x)
            else:
                x_data = []
            
            if hasattr(trace, 'y') and trace.y is not None:
                y_data = list(trace.y)
            else:
                y_data = []
            
            # データをワークシートに書き込む
            start_col = col_offset + 1
            if trace_idx == 0:
                # 最初の系列のみヘッダー行にグラフタイトルを配置
                ws.cell(row=2, column=start_col, value=f"{series_name} (X)")
                ws.cell(row=2, column=start_col + 1, value=f"{series_name} (Y)")
            else:
                ws.cell(row=2, column=start_col, value=f"{series_name} (X)")
                ws.cell(row=2, column=start_col + 1, value=f"{series_name} (Y)")
            
            for row_idx, (x_val, y_val) in enumerate(zip(x_data, y_data), start=3):
                ws.cell(row=row_idx, column=start_col, value=x_val)
                ws.cell(row=row_idx, column=start_col + 1, value=y_val)
            
            col_offset += 2
        
        # 散布図を作成
        chart = ScatterChart()
        chart.style = None  # プレーンな設定
        if graph_title:
            chart.title = graph_title
        chart.y_axis.title = fig.layout.yaxis.title.text if fig.layout.yaxis and fig.layout.yaxis.title else ""
        chart.x_axis.title = fig.layout.xaxis.title.text if fig.layout.xaxis and fig.layout.xaxis.title else ""
        
        # 各系列を追加
        for trace_idx in range(len(fig.data)):
            start_col = trace_idx * 2 + 1
            max_row = len(list(fig.data[trace_idx].y)) + 2 if hasattr(fig.data[trace_idx], 'y') else 3
            
            xvalues = Reference(ws, min_col=start_col, min_row=3, max_row=max_row)
            yvalues = Reference(ws, min_col=start_col + 1, min_row=3, max_row=max_row)
            series = chart.series.append(yvalues)
            series.xvalues = xvalues
            series.title = ws.cell(row=2, column=start_col + 1).value
        
        # グラフのサイズを設定
        chart.width = 15
        chart.height = 10
        
        # 凡例の設定
        if len(series_names) > 1:
            chart.legend.position = 'r'
        else:
            chart.legend = None
        
        # グラフの配置
        max_data_row = max([len(list(trace.y)) for trace in fig.data if hasattr(trace, 'y')]) if fig.data else 3
        ws.add_chart(chart, "A" + str(max_data_row + 4))
    
    else:
        # その他のグラフタイプ（デフォルトは折れ線グラフとして処理）
        # X軸のカテゴリを取得（ticktextを優先的に使用）
        categories = []

        # まずticktextを確認
        try:
            if hasattr(fig.layout, 'xaxis') and fig.layout.xaxis is not None:
                if hasattr(fig.layout.xaxis, 'ticktext') and fig.layout.xaxis.ticktext is not None:
                    ticktext = fig.layout.xaxis.ticktext
                    if isinstance(ticktext, (list, tuple)) and len(ticktext) > 0:
                        categories = [str(t) for t in ticktext]
        except Exception as e:
            pass

        # ticktextが取得できなかった場合のみ、xの値を使用
        if not categories:
            try:
                if fig.data and hasattr(fig.data[0], 'x') and fig.data[0].x is not None:
                    x_data = fig.data[0].x
                    if isinstance(x_data, (list, tuple)) and len(x_data) > 0:
                        categories = [str(x) for x in x_data]
            except Exception as e:
                categories = []
        
        # データを収集
        for trace in fig.data:
            if hasattr(trace, 'name') and trace.name:
                series_name = trace.name
            else:
                series_name = f"系列{len(data_dict) + 1}"
            
            series_names.append(series_name)
            
            if hasattr(trace, 'y') and trace.y is not None:
                data_dict[series_name] = list(trace.y)
        
        if data_dict:  # データがある場合のみ処理
            # データをワークシートに書き込む
            ws.cell(row=1, column=1, value="カテゴリ")
            col_idx = 2
            for series_name in series_names:
                ws.cell(row=1, column=col_idx, value=series_name)
                col_idx += 1
            # グラフタイトルを1行目の最後の列の次に配置
            ws.cell(row=1, column=col_idx, value=graph_title)
            
            for row_idx, category in enumerate(categories, start=2):
                ws.cell(row=row_idx, column=1, value=category)
                col_idx = 2
                for series_name in series_names:
                    if row_idx - 2 < len(data_dict[series_name]):
                        ws.cell(row=row_idx, column=col_idx, value=data_dict[series_name][row_idx - 2])
                    col_idx += 1
            
            # 折れ線グラフを作成
            chart = LineChart()
            chart.style = None  # プレーンな設定
            if graph_title:
                chart.title = graph_title
            chart.y_axis.title = fig.layout.yaxis.title.text if fig.layout.yaxis and fig.layout.yaxis.title else ""
            chart.x_axis.title = fig.layout.xaxis.title.text if fig.layout.xaxis and fig.layout.xaxis.title else ""
            
            # データ範囲を設定
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(categories) + 1, max_col=len(series_names) + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(categories) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            # グラフのサイズを設定
            chart.width = 15
            chart.height = 10
            
            # 凡例の設定
            if len(series_names) > 1:
                chart.legend.position = 'r'
            else:
                chart.legend = None  # 凡例を非表示
            
            # グラフの配置
            ws.add_chart(chart, "A" + str(len(categories) + 3))
    
    # ExcelファイルをBytesIOに保存
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()
